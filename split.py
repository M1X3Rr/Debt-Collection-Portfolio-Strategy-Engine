import json
import os
import random
from typing import Dict, Tuple

import joblib
import numpy as np
import pandas as pd
from sklearn.base import BaseEstimator, TransformerMixin
from sklearn.model_selection import train_test_split
from sklearn.preprocessing import OrdinalEncoder, StandardScaler


def _normalize_product_placeholder(df: pd.DataFrame, config: dict) -> pd.DataFrame:
    """
    Replace invalid/missing Product values with a stable placeholder.

    If a dedicated debtor name column exists (default: "Name") and Product is
    missing/blank or equals the debtor name, treat Product as missing and set
    it to the configured placeholder (default: "Unknown_Product").
    """
    cols_cfg = config.get("columns", {})
    product_col = cols_cfg.get("product_column", "Product")
    name_col = cols_cfg.get("debtor_name_column", "Name")
    placeholder = cols_cfg.get("product_missing_placeholder", "Unknown_Product")

    if product_col not in df.columns:
        return df

    out = df.copy()
    prod = out[product_col]
    prod_str = prod.astype(str).str.strip()

    missing_mask = prod.isna() | (prod_str == "") | (prod_str.str.lower() == "nan")

    if name_col in out.columns:
        name_str = out[name_col].astype(str).str.strip()
        missing_mask = missing_mask | (prod_str == name_str)

    if missing_mask.any():
        out.loc[missing_mask, product_col] = placeholder

    return out


class LogStandardScaler(BaseEstimator, TransformerMixin):
    """
    A scaler that applies log transformation before standardization.
    
    This helps with skewed numerical features (like debt values) by:
    1. Applying log1p(x) = log(1 + x) to compress the range
    2. Then applying standard scaling (mean=0, std=1)
    
    The log transformation converts multiplicative relationships to additive ones,
    making it easier for neural networks to learn patterns in highly skewed data.
    
    Example:
        - 50 EUR debt -> log1p(50) ≈ 3.93
        - 50,000 EUR debt -> log1p(50000) ≈ 10.82
        
    This compressed range (3.93 to 10.82) is much easier for the model to learn
    than the original range (50 to 50,000).
    """
    
    def __init__(self, log_features: list = None):
        """
        Parameters
        ----------
        log_features : list, optional
            List of feature names that should have log transformation applied.
            If None, log transform is applied to all features.
        """
        self.log_features = log_features
        self.standard_scaler = StandardScaler()
        self.feature_names_in_ = None
        self._log_mask = None
    
    def fit(self, X, y=None):
        """Fit the scaler on training data."""
        if hasattr(X, 'columns'):
            self.feature_names_in_ = list(X.columns)
            X = X.values
        else:
            self.feature_names_in_ = [f"feature_{i}" for i in range(X.shape[1])]
        
        # Determine which features to log-transform
        if self.log_features is not None:
            self._log_mask = np.array([
                name in self.log_features 
                for name in self.feature_names_in_
            ])
        else:
            # By default, apply log to all features
            self._log_mask = np.ones(X.shape[1], dtype=bool)
        
        # Apply log transformation to selected features
        X_transformed = X.copy().astype(np.float64)
        X_transformed[:, self._log_mask] = np.log1p(
            np.clip(X_transformed[:, self._log_mask], 0, None)
        )
        
        # Fit the standard scaler on log-transformed data
        self.standard_scaler.fit(X_transformed)
        
        # Store statistics for reference
        self.mean_ = self.standard_scaler.mean_
        self.scale_ = self.standard_scaler.scale_
        
        return self
    
    def transform(self, X):
        """Transform features using log + standard scaling."""
        if hasattr(X, 'values'):
            X = X.values
        
        X_transformed = X.copy().astype(np.float64)
        
        # Apply log transformation to selected features
        X_transformed[:, self._log_mask] = np.log1p(
            np.clip(X_transformed[:, self._log_mask], 0, None)
        )
        
        # Apply standard scaling
        return self.standard_scaler.transform(X_transformed)
    
    def inverse_transform(self, X):
        """Reverse the transformation."""
        # First, reverse standard scaling
        X_inv = self.standard_scaler.inverse_transform(X)
        
        # Then, reverse log transformation (expm1 = exp(x) - 1)
        # Clip values to prevent overflow and replace any inf / NaN with a
        # large but finite cap so downstream models don't explode numerically.
        log_values = X_inv[:, self._log_mask]
        log_values_clipped = np.clip(log_values, None, 700)
        inv_vals = np.expm1(log_values_clipped)
        # Replace non‑finite values with a conservative cap.
        cap = 1e9
        inv_vals = np.where(np.isfinite(inv_vals), inv_vals, np.sign(inv_vals) * cap)
        X_inv[:, self._log_mask] = inv_vals
        
        return X_inv
    
    def fit_transform(self, X, y=None):
        """Fit and transform in one step."""
        return self.fit(X, y).transform(X)


def load_config(config_path: str = "config.json") -> dict:
    """
    Load project configuration from a JSON file.

    Parameters
    ----------
    config_path : str
        Path to the configuration JSON file.

    Returns
    -------
    dict
        Parsed configuration dictionary.
    """
    with open(config_path, "r", encoding="utf-8") as f:
        return json.load(f)


def load_raw_data(path: str) -> pd.DataFrame:
    """
    Load the raw portfolio data from a file (Excel or CSV).

    Parameters
    ----------
    path : str
        Path to the input data file. Supports `.xlsx`, `.xls`, and `.csv`.

    Returns
    -------
    pd.DataFrame
        Loaded raw data.
    """
    #NOTE: This helper automatically chooses between Excel and CSV
    # based on the file extension, so you can switch formats via config only.
    lower = path.lower()
    if lower.endswith((".xlsx", ".xls")):
        return pd.read_excel(path)
    try:
        return pd.read_csv(path, encoding="utf-8")
    except (UnicodeDecodeError, UnicodeError):
        return pd.read_csv(path, encoding="cp1252")


def load_table(path: str) -> pd.DataFrame:
    """
    Generic loader for train/valid/test files (CSV or Excel).

    Parameters
    ----------
    path : str
        Path to the data file. Supports `.csv`, `.xlsx`, `.xls`.

    Returns
    -------
    pd.DataFrame
        Loaded DataFrame.
    """
    lower = path.lower()
    if lower.endswith((".xlsx", ".xls")):
        return pd.read_excel(path)
    
    if lower.endswith(".csv"):
        try:
            return pd.read_csv(path, encoding="utf-8")
        except (UnicodeDecodeError, UnicodeError):
            return pd.read_csv(path, encoding="cp1252")
            
    raise ValueError(f"Unsupported file format for {path}. Use .csv, .xlsx or .xls.")


def _convert_datetime_to_age(df: pd.DataFrame, col: str) -> pd.Series:
    """
    Convert a datetime column (e.g. birthdate) to numeric age in years.
    """
    if col not in df.columns:
        return pd.Series(dtype=float)

    series = df[col]

    # Direct datetime dtype.
    if pd.api.types.is_datetime64_any_dtype(series):
        today = pd.Timestamp.now()
        ages = (today - series).dt.days / 365.25
        return ages.round().astype(float)

    # Object that can be parsed as datetime for the majority of rows.
    if series.dtype == object:
        try:
            parsed = pd.to_datetime(series, errors="coerce")
            if parsed.notna().sum() > len(series) * 0.5:
                today = pd.Timestamp.now()
                ages = (today - parsed).dt.days / 365.25
                return ages.round().astype(float)
        except Exception:
            pass

    # Numeric values that may actually be timestamps (very large magnitudes).
    numeric_series = pd.to_numeric(series, errors="coerce")
    if numeric_series.notna().any():
        if numeric_series.abs().max() > 1e10:
            try:
                dt_series = pd.to_datetime(numeric_series, unit="ns", errors="coerce")
                if dt_series.notna().any():
                    today = pd.Timestamp.now()
                    ages = (today - dt_series).dt.days / 365.25
                    return ages.clip(0, 120).astype(float)
            except Exception:
                # Fall through to plain numeric coercion below.
                pass

    return pd.to_numeric(series, errors="coerce")


def get_target_series(df: pd.DataFrame, config: dict) -> pd.Series:
    """
    Return the training target series according to target_transform config.

    When mode == 'paid_ratio', this returns Paid Value / Case Value clipped
    to non-negative values. Otherwise it returns the raw Paid Value column.
    """
    cols_cfg = config["columns"]
    target_col = cols_cfg["target_column"]

    if target_col not in df.columns:
        return pd.Series(dtype=float)

    base = pd.to_numeric(df[target_col], errors="coerce")

    tt_cfg = config.get("target_transform", {})
    mode = str(tt_cfg.get("mode", "paid_value")).lower()
    if mode != "paid_ratio":
        return base

    case_value_col = "Case Value"
    if case_value_col not in df.columns:
        return base

    case_values = pd.to_numeric(df[case_value_col], errors="coerce")
    ratio = base / case_values.replace(0, np.nan)
    ratio = ratio.replace([np.inf, -np.inf], np.nan)
    ratio = ratio.clip(lower=0)
    return ratio


def case_duration_weeks_series(df: pd.DataFrame, config: dict) -> pd.Series:
    """
    Weeks each case was active (End Date - Import Date).

    Returns NaN for rows with missing/invalid dates or non-positive duration so
    callers can exclude them from weekly-rate averages (avoids treating totals
    as \"per week\" when duration is unknown).
    """
    cols_cfg = config["columns"]
    date_cfg = cols_cfg.get("date_features", {})
    import_col = date_cfg.get("import_date", "Import date")
    end_col = date_cfg.get("end_date", "End Date")
    if import_col not in df.columns or end_col not in df.columns:
        return pd.Series(np.nan, index=df.index)

    import_dates = pd.to_datetime(df[import_col], errors="coerce")
    end_dates = pd.to_datetime(df[end_col], errors="coerce")
    days_diff = (end_dates - import_dates).dt.days
    valid = import_dates.notna() & end_dates.notna() & (days_diff > 0)
    weeks = pd.Series(np.nan, index=df.index, dtype=float)
    weeks.loc[valid] = days_diff.loc[valid] / 7.0
    return weeks


def compute_weekly_rates(df: pd.DataFrame, config: dict) -> pd.DataFrame:
    """
    Calculate weekly action rates based on case duration (End Date - Import Date).

    For each action feature (Calls, Letters, SMS), creates a new column
    with the suffix '_per_week' containing the weekly rate.

    Parameters
    ----------
    df : pd.DataFrame
        Input DataFrame containing date and action columns.
    config : dict
        Configuration dictionary with date_features and action_features.

    Returns
    -------
    pd.DataFrame
        DataFrame with additional weekly rate columns.
    """
    cols_cfg = config["columns"]
    date_cfg = cols_cfg.get("date_features", {})
    import_col = date_cfg.get("import_date", "Import date")
    end_col = date_cfg.get("end_date", "End Date")
    action_features = cols_cfg.get("action_features", [])

    # Check if required date columns exist
    if import_col not in df.columns or end_col not in df.columns:
        print(f"Warning: Date columns '{import_col}' or '{end_col}' not found. Skipping weekly rate calculation.")
        return df

    df = df.copy()
    weeks = case_duration_weeks_series(df, config)

    # NOTE: Duration normalization corrects cumulative bias. Structural features
    # (Debtor Age, Location) remain in the feature set and are not weakened.
    # Calculate weekly rates for each action (NaN when duration invalid).
    for action in action_features:
        if action in df.columns:
            weekly_col = f"{action}_per_week"
            action_values = pd.to_numeric(df[action], errors="coerce").fillna(0)
            rates = action_values / weeks
            df[weekly_col] = rates.where(weeks.notna()).round(4)
        else:
            print(f"Warning: Action column '{action}' not found.")

    n_valid = int(weeks.notna().sum())
    print(
        f"Computed weekly rates for {len(action_features)} action features "
        f"({n_valid}/{len(df)} rows with valid Import/End duration)."
    )
    return df


def filter_valid_training_data(df: pd.DataFrame, config: dict) -> pd.DataFrame:
    """
    Filter out cases with invalid Paid Value (null, zero, or negative).

    Parameters
    ----------
    df : pd.DataFrame
        Input DataFrame.
    config : dict
        Configuration dictionary with data_validation settings.

    Returns
    -------
    pd.DataFrame
        Filtered DataFrame with only valid training samples.
    """
    cols_cfg = config["columns"]
    target_col = cols_cfg["target_column"]
    validation = config.get("data_validation", {})

    if target_col not in df.columns:
        print(f"Warning: Target column '{target_col}' not found. Skipping validation filter.")
        return df

    original_count = len(df)
    df = df.copy()

    # Convert target to numeric
    df[target_col] = pd.to_numeric(df[target_col], errors="coerce")

    # Remove null values
    df = df.dropna(subset=[target_col])
    null_removed = original_count - len(df)

    # Remove zero values if configured.
    # NOTE: Excluding zero-paid cases may make the model more optimistic; the distribution of non-performers matters for forecasting.
    zero_removed = 0
    if validation.get("exclude_zero_paid", True):
        before = len(df)
        df = df[df[target_col] != 0]
        zero_removed = before - len(df)

    # Remove negative values if configured
    negative_removed = 0
    if validation.get("exclude_negative_paid", True):
        before = len(df)
        df = df[df[target_col] > 0]
        negative_removed = before - len(df)

    # Optionally cap Paid Value at a multiple of Case Value for training only,
    # so extreme outliers do not dominate the target distribution. This keeps
    # the model calibrated while still allowing predictions above standard
    # recovery when justified.
    exceeds_case_value_removed = 0
    case_value_col = "Case Value"
    if case_value_col in df.columns:
        case_values = pd.to_numeric(df[case_value_col], errors="coerce").fillna(0)
        paid_values = pd.to_numeric(df[target_col], errors="coerce")

        tt_cfg = config.get("target_transform", {})
        max_training_ratio = tt_cfg.get("max_training_ratio")
        if max_training_ratio is not None:
            max_training_ratio = float(max_training_ratio)
            max_allowed_training = case_values * max_training_ratio
            capped_paid = np.minimum(paid_values, max_allowed_training)
            n_capped = int((capped_paid < paid_values).sum())
            if n_capped > 0:
                print(
                    f"Data validation: Capped {n_capped} rows where Paid Value exceeded "
                    f"{max_training_ratio:.2f}x Case Value for training stability."
                )
            df[target_col] = capped_paid

        # Optional strict filtering of rows still exceeding a small tolerance
        # above Case Value (e.g. data errors). This operates on the (possibly
        # capped) Paid Value values.
        if validation.get("exclude_exceeds_case_value", True):
            before = len(df)
            paid_values = pd.to_numeric(df[target_col], errors="coerce")
            tolerance = validation.get("case_value_tolerance", 0.01)
            max_allowed = case_values * (1 + tolerance)
            df = df[paid_values <= max_allowed]
            exceeds_case_value_removed = before - len(df)
            if exceeds_case_value_removed > 0:
                print(
                    f"  Filtered {exceeds_case_value_removed} cases where Paid Value exceeds "
                    f"Case Value beyond tolerance (possible data quality issues)."
                )

    # Remove cases where Case Value is 0, null, or negative.
    case_value_removed = 0
    if validation.get("exclude_zero_or_negative_case_value", True):
        case_value_col = "Case Value"
        if case_value_col in df.columns:
            before = len(df)
            case_values = pd.to_numeric(df[case_value_col], errors="coerce")
            df = df[case_values.notna() & (case_values > 0)]
            case_value_removed = before - len(df)

    # Remove cases where all action columns are 0, null, or negative (no contact effort).
    action_cols = cols_cfg.get("action_features", [])
    all_actions_zero_removed = 0
    if validation.get("exclude_all_actions_zero", True) and action_cols:
        present = [c for c in action_cols if c in df.columns]
        if present:
            before = len(df)
            action_sums = pd.Series(0, index=df.index)
            for c in present:
                action_sums += pd.to_numeric(df[c], errors="coerce").fillna(0)
            df = df[action_sums > 0]
            all_actions_zero_removed = before - len(df)

    total_filtered = original_count - len(df)
    print(f"Data validation: Filtered {total_filtered} rows "
          f"(null: {null_removed}, zero: {zero_removed}, negative: {negative_removed}, "
          f"exceeds_case_value: {exceeds_case_value_removed}, case_value_invalid: {case_value_removed}, all_actions_zero: {all_actions_zero_removed})")
    print(f"Remaining valid samples: {len(df)}")

    return df


def basic_cleaning(df: pd.DataFrame, config: dict) -> pd.DataFrame:
    """
    Apply basic data cleaning: drop rows with missing target, fill missing
    numericals with median and categoricals with 'Unknown'.

    Parameters
    ----------
    df : pd.DataFrame
        Input raw DataFrame.
    config : dict
        Project configuration containing feature definitions.

    Returns
    -------
    pd.DataFrame
        Cleaned DataFrame.
    """
    cols_cfg = config["columns"]
    target_col = cols_cfg["target_column"]

    # Drop rows without target to avoid training on incomplete samples.
    df = df.dropna(subset=[target_col]).copy()

    # Normalize Product when it was filled with debtor Name.
    df = _normalize_product_placeholder(df, config)

    categorical_features = cols_cfg["categorical_features"]
    numerical_features = cols_cfg["numerical_features"]
    action_features = cols_cfg["action_features"]

    # Convert datetime columns to numeric (e.g. birthdate -> age).
    for col in numerical_features:
        if col in df.columns:
            if pd.api.types.is_datetime64_any_dtype(df[col]) or (
                df[col].dtype == object
                and df[col].dropna().apply(lambda x: isinstance(x, pd.Timestamp) or hasattr(x, "year")).any()
            ):
                df.loc[:, col] = _convert_datetime_to_age(df, col)

    # Fill numericals (including action features which are numeric controls).
    for col in numerical_features + action_features:
        if col in df.columns:
            df.loc[:, col] = pd.to_numeric(df[col], errors="coerce")
            median_val = df[col].median()
            df.loc[:, col] = df[col].fillna(median_val if pd.notna(median_val) else 0)

    # Fill categoricals.
    for col in categorical_features:
        if col in df.columns:
            df.loc[:, col] = df[col].fillna("Unknown")

    return df


def _compute_fill_values(train_df: pd.DataFrame, config: dict) -> Dict[str, float]:
    cols_cfg = config["columns"]
    target_col = cols_cfg["target_column"]
    train_df = train_df.dropna(subset=[target_col]).copy()
    numerical_features = cols_cfg["numerical_features"]
    action_features = cols_cfg["action_features"]
    weekly_action_features = cols_cfg.get("weekly_action_features", [])

    # Ensure numerical features are numeric; do NOT try to interpret them
    # as datetimes here. Dedicated date columns are handled separately.
    for col in numerical_features:
        if col in train_df.columns:
            train_df.loc[:, col] = pd.to_numeric(train_df[col], errors="coerce")

    fill_values: Dict[str, float] = {}
    # Include weekly features in fill values computation
    all_numeric_cols = numerical_features + action_features + weekly_action_features
    for col in all_numeric_cols:
        if col in train_df.columns:
            median_val = train_df[col].median()
            fill_values[col] = float(median_val) if pd.notna(median_val) else 0.0
    return fill_values


def _apply_cleaning(
    df: pd.DataFrame,
    config: dict,
    numeric_fill_values: Dict[str, float],
) -> pd.DataFrame:
    cols_cfg = config["columns"]
    target_col = cols_cfg["target_column"]
    categorical_features = cols_cfg["categorical_features"]
    numerical_features = cols_cfg["numerical_features"]
    action_features = cols_cfg["action_features"]
    weekly_action_features = cols_cfg.get("weekly_action_features", [])

    df = df.dropna(subset=[target_col]).copy()

    # Normalize Product when it was filled with debtor Name.
    df = _normalize_product_placeholder(df, config)

    # Ensure numerical features are numeric; do NOT try to interpret them
    # as datetimes here. Dedicated date columns are handled separately.
    for col in numerical_features:
        if col in df.columns:
            df.loc[:, col] = pd.to_numeric(df[col], errors="coerce")

    # Include weekly features in numeric processing
    all_numeric_cols = numerical_features + action_features + weekly_action_features
    for col in all_numeric_cols:
        if col in df.columns:
            df.loc[:, col] = pd.to_numeric(df[col], errors="coerce")
            df.loc[:, col] = df[col].fillna(numeric_fill_values.get(col, 0))

    for col in categorical_features:
        if col in df.columns:
            df.loc[:, col] = df[col].fillna("Unknown")

    return df


def _ensure_categorical_columns(df: pd.DataFrame, config: dict) -> pd.DataFrame:
    """Add any configured categorical columns that are missing (e.g. Product in legacy data)."""
    cols_cfg = config["columns"]
    categorical_features = [c for c in cols_cfg["categorical_features"] if c not in cols_cfg.get("exclude_features", [])]
    missing = [c for c in categorical_features if c not in df.columns]
    if not missing:
        return df
    df = df.copy()
    for col in missing:
        df[col] = "Unknown"
    return df


def prepare_splits(
    train_df: pd.DataFrame,
    valid_df: pd.DataFrame,
    test_df: pd.DataFrame,
    config: dict,
) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """
    Apply consistent cleaning across splits using train statistics.
    Also computes weekly action rates and filters invalid training data.

    Date columns named in config ``columns.date_features`` remain on the exported
    splits when present in the source file; GBM training/inference uses them to
    derive extended features (for example ``days_active`` and import batch id).
    """
    # First, filter invalid training data (Paid Value <= 0 or null)
    train_df = filter_valid_training_data(train_df, config)
    valid_df = filter_valid_training_data(valid_df, config)
    test_df = filter_valid_training_data(test_df, config)

    # Ensure all configured categorical columns exist (e.g. Product) so encoder and cleaning succeed
    train_df = _ensure_categorical_columns(train_df, config)
    valid_df = _ensure_categorical_columns(valid_df, config)
    test_df = _ensure_categorical_columns(test_df, config)

    # Compute weekly rates for each split
    train_df = compute_weekly_rates(train_df, config)
    valid_df = compute_weekly_rates(valid_df, config)
    test_df = compute_weekly_rates(test_df, config)

    # Apply standard cleaning
    fill_values = _compute_fill_values(train_df, config)
    train_df = _apply_cleaning(train_df, config, fill_values)
    valid_df = _apply_cleaning(valid_df, config, fill_values)
    test_df = _apply_cleaning(test_df, config, fill_values)

    return train_df, valid_df, test_df


def stream_split_excel(input_path: str, config: dict) -> None:
    """
    Stream a large Excel file and split it into train/valid/test without
    loading the entire dataset into memory. Splits by row at random.
    For split by (Client, Import date) use the non-streaming path instead
    (load raw data, then split_dataset).
    """
    from openpyxl import Workbook, load_workbook

    data_cfg = config["data"]
    train_cfg = config["training"]

    os.makedirs(os.path.dirname(data_cfg["train_path"]), exist_ok=True)
    os.makedirs(os.path.dirname(data_cfg["valid_path"]), exist_ok=True)
    os.makedirs(os.path.dirname(data_cfg["test_path"]), exist_ok=True)

    wb = load_workbook(input_path, read_only=True, data_only=True)
    ws = wb.active

    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if header_row is None:
        raise ValueError("Input Excel file is empty.")

    rng = random.Random(train_cfg.get("random_state", 42))
    test_size = float(train_cfg.get("test_split", 0.15))
    val_size = float(train_cfg.get("val_split", 0.15))

    train_wb = Workbook(write_only=True)
    valid_wb = Workbook(write_only=True)
    test_wb = Workbook(write_only=True)

    train_ws = train_wb.create_sheet("train")
    valid_ws = valid_wb.create_sheet("valid")
    test_ws = test_wb.create_sheet("test")

    train_ws.append(list(header_row))
    valid_ws.append(list(header_row))
    test_ws.append(list(header_row))

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row is None or all(cell is None for cell in row):
            continue
        r = rng.random()
        if r < test_size:
            test_ws.append(list(row))
        elif r < test_size + val_size:
            valid_ws.append(list(row))
        else:
            train_ws.append(list(row))

    train_wb.save(data_cfg["train_path"])
    valid_wb.save(data_cfg["valid_path"])
    test_wb.save(data_cfg["test_path"])


def split_dataset(
    df: pd.DataFrame,
    config: dict,
) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """
    Split the cleaned DataFrame into train, validation, and test sets.

    When split_by_import is True (and Client + Import date columns exist),
    splits by (Client, Import date) groups so that all cases from the same
    import stay in the same split. This lets the model see coherent
    portfolio behaviour per client and import.

    Parameters
    ----------
    df : pd.DataFrame
        Cleaned DataFrame.
    config : dict
        Configuration dictionary including split ratios and random_state.

    Returns
    -------
    Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]
        Train, validation, and test DataFrames.
    """
    train_cfg = config["training"]
    test_size = train_cfg["test_split"]
    val_size = train_cfg["val_split"]
    random_state = train_cfg["random_state"]
    split_by_import = train_cfg.get("split_by_import", False)

    cols_cfg = config["columns"]
    target_col = cols_cfg["target_column"]
    date_cfg = cols_cfg.get("date_features", {})
    import_col = date_cfg.get("import_date", "Import date")
    client_col = "Client"

    df = df.reset_index(drop=True)

    # Split by (Client, Import date) groups when requested and columns exist
    if split_by_import and client_col in df.columns and import_col in df.columns:
        # Build a stable group key (string) per row
        group_key = (
            df[client_col].astype(str)
            + "\n"
            + pd.to_datetime(df[import_col], errors="coerce").astype(str)
        )
        df = df.copy()
        df["_import_group_"] = group_key
        unique_groups = group_key.unique()
        rng = np.random.RandomState(random_state)
        rng.shuffle(unique_groups)
        n = len(unique_groups)
        n_test = max(1, int(n * test_size))
        n_val = max(0, int(n * val_size))
        n_train = n - n_test - n_val
        if n_train < 1:
            n_train, n_test, n_val = n - 2, 1, 1
        train_groups = set(unique_groups[:n_train])
        val_groups = set(unique_groups[n_train : n_train + n_val])
        test_groups = set(unique_groups[n_train + n_val :])
        train_df = df[df["_import_group_"].isin(train_groups)].drop(columns=["_import_group_"])
        valid_df = df[df["_import_group_"].isin(val_groups)].drop(columns=["_import_group_"])
        test_df = df[df["_import_group_"].isin(test_groups)].drop(columns=["_import_group_"])
        print(
            f"Split by import: {len(train_groups)} train, {len(val_groups)} valid, {len(test_groups)} test groups; "
            f"rows: {len(train_df)}, {len(valid_df)}, {len(test_df)}"
        )
        return train_df.reset_index(drop=True), valid_df.reset_index(drop=True), test_df.reset_index(drop=True)

    # Row-wise split (original behaviour)
    feature_cols = (
        cols_cfg["categorical_features"]
        + cols_cfg["numerical_features"]
        + cols_cfg["action_features"]
    )
    exclude_cols = set(cols_cfg.get("exclude_features", []))
    feature_cols = [c for c in feature_cols if c not in exclude_cols and c != target_col]

    train_valid_df, test_df = train_test_split(
        df, test_size=test_size, random_state=random_state, shuffle=True
    )
    relative_val_size = val_size / (1.0 - test_size)
    train_df, valid_df = train_test_split(
        train_valid_df,
        test_size=relative_val_size,
        random_state=random_state,
        shuffle=True,
    )
    return train_df, valid_df, test_df


def fit_preprocessors(
    train_df: pd.DataFrame,
    config: dict,
) -> Tuple[OrdinalEncoder, LogStandardScaler, LogStandardScaler]:
    """
    Fit categorical encoder and numerical scaler on the training data only.
    
    Uses LogStandardScaler which applies log1p transformation before standard
    scaling to handle skewed distributions (e.g., debt values ranging from
    50 to 50,000 EUR). This makes it easier for the neural network to learn.

    Parameters
    ----------
    train_df : pd.DataFrame
        Training DataFrame.
    config : dict
        Configuration dictionary containing feature definitions.

    Returns
    -------
    Tuple[OrdinalEncoder, LogStandardScaler, LogStandardScaler]
        Fitted (categorical encoder, feature scaler, target scaler).
    """
    cols_cfg = config["columns"]
    categorical_features = cols_cfg["categorical_features"]
    numerical_features = cols_cfg["numerical_features"]
    action_features = cols_cfg["action_features"]
    weekly_action_features = cols_cfg.get("weekly_action_features", [])
    exclude_cols = set(cols_cfg.get("exclude_features", []))

    # Filter features with the same logic used for splitting.
    categorical_features = [c for c in categorical_features if c not in exclude_cols]

    # Include weekly action features in numeric features if they exist in the dataframe
    numeric_like_features = [
        c
        for c in (numerical_features + action_features)
        if c not in exclude_cols
    ]

    # Add weekly features if they exist in the training data
    for wf in weekly_action_features:
        if wf in train_df.columns and wf not in numeric_like_features:
            numeric_like_features.append(wf)

    # OrdinalEncoder: map categories to integer indices used by embeddings.
    # NOTE: handle_unknown="use_encoded_value" keeps inference robust to new categories.
    encoder = OrdinalEncoder(
        handle_unknown="use_encoded_value",
        unknown_value=-1,
        dtype=np.int64,
    )
    if categorical_features:
        encoder.fit(train_df[categorical_features])

    # Determine which features should be log-transformed
    # Apply log to monetary values and counts (skewed distributions)
    log_features = []
    for col in numeric_like_features:
        if col in train_df.columns:
            # Log-transform monetary values (Case Value) and action counts
            if col in ["Case Value"] + action_features + weekly_action_features:
                log_features.append(col)
            # Also log-transform any feature with high skewness
            elif train_df[col].skew() > 1.0:
                log_features.append(col)
    
    print(f"Log-transforming features: {log_features}")
    
    # Use LogStandardScaler for features
    scaler = LogStandardScaler(log_features=log_features if log_features else None)
    # Only fit on columns that actually exist
    available_numeric = [c for c in numeric_like_features if c in train_df.columns]
    if available_numeric:
        scaler.fit(train_df[available_numeric])

    # Scale target using LogStandardScaler to handle skewed distributions.
    # When target_transform.mode == 'paid_ratio', the scaler is fit on the
    # Paid/Case ratio; otherwise it is fit on raw Paid Value.
    target_series = get_target_series(train_df, config)
    target_series = target_series.fillna(0.0)
    target_df = pd.DataFrame({"target": target_series.astype(float)})

    target_scaler = LogStandardScaler(log_features=["target"])
    target_scaler.fit(target_df[["target"]])

    print("Target statistics before log transform (core space):")
    print(f"  Mean: {target_df['target'].mean():.4f}, Std: {target_df['target'].std():.4f}")
    print(f"  Min: {target_df['target'].min():.4f}, Max: {target_df['target'].max():.4f}")
    print(f"  Skewness: {target_df['target'].skew():.4f}")

    return encoder, scaler, target_scaler


def save_splits_and_artifacts(
    train_df: pd.DataFrame,
    valid_df: pd.DataFrame,
    test_df: pd.DataFrame,
    encoder: OrdinalEncoder,
    scaler: LogStandardScaler,
    target_scaler: LogStandardScaler,
    config: dict,
) -> None:
    """
    Save the dataset splits and preprocessing artifacts to disk.

    Parameters
    ----------
    train_df : pd.DataFrame
        Training DataFrame.
    valid_df : pd.DataFrame
        Validation DataFrame.
    test_df : pd.DataFrame
        Test DataFrame.
    encoder : OrdinalEncoder
        Fitted categorical encoder.
    scaler : LogStandardScaler
        Fitted numerical scaler with log transformation.
    target_scaler : LogStandardScaler
        Fitted target scaler with log transformation.
    config : dict
        Configuration dictionary with paths.
    """
    data_cfg = config["data"]
    artifacts_dir = data_cfg["artifacts_dir"]

    # Ensure output directories exist.
    os.makedirs(os.path.dirname(data_cfg["train_path"]), exist_ok=True)
    os.makedirs(os.path.dirname(data_cfg["valid_path"]), exist_ok=True)
    os.makedirs(os.path.dirname(data_cfg["test_path"]), exist_ok=True)
    os.makedirs(artifacts_dir, exist_ok=True)

    # Save splits using a format inferred from the file extension.
    def _save(df: pd.DataFrame, path: str) -> None:
        lower = path.lower()
        if lower.endswith(".csv"):
            df.to_csv(path, index=False)
        elif lower.endswith((".xlsx", ".xls")):
            df.to_excel(path, index=False)
        else:
            # Default to CSV if extension is unknown.
            df.to_csv(path, index=False)

    _save(train_df, data_cfg["train_path"])
    _save(valid_df, data_cfg["valid_path"])
    _save(test_df, data_cfg["test_path"])

    # Save preprocessing artifacts.
    joblib.dump(
        encoder,
        os.path.join(artifacts_dir, "encoder.pkl"),
    )
    joblib.dump(
        scaler,
        os.path.join(artifacts_dir, "scaler.pkl"),
    )
    joblib.dump(
        target_scaler,
        os.path.join(artifacts_dir, "target_scaler.pkl"),
    )


def main() -> None:
    """
    Main entry point for data preparation and splitting.

    This will:
    1. Load configuration.
    2. Load raw data.
    3. Clean the data.
    4. Split into train/validation/test sets.
    5. Fit encoder/scaler on the training set.
    6. Persist the splits and preprocessing artifacts to disk.
    """
    config = load_config()
    data_cfg = config["data"]

    train_path = data_cfg["train_path"]
    valid_path = data_cfg["valid_path"]
    test_path = data_cfg["test_path"]

    # If explicit train/valid/test files already exist (e.g. train.xlsx, etc.),
    # use them directly to fit the preprocessors and only (re)save artifacts.
    if all(os.path.exists(p) for p in (train_path, valid_path, test_path)):
        train_df = load_table(train_path)
        valid_df = load_table(valid_path)
        test_df = load_table(test_path)
    else:
        # Fallback: create splits from a raw labeled dataset pointed to by input_path.
        input_path = data_cfg["input_path"]
        if input_path.lower().endswith(".xlsx"):
            stream_split_excel(input_path, config)
            train_df = load_table(train_path)
            valid_df = load_table(valid_path)
            test_df = load_table(test_path)
        else:
            raw_df = load_raw_data(input_path)
            cleaned_df = basic_cleaning(raw_df, config)
            train_df, valid_df, test_df = split_dataset(cleaned_df, config)

    # Apply consistent cleaning across splits using train medians.
    train_df, valid_df, test_df = prepare_splits(train_df, valid_df, test_df, config)

    # Fit preprocessing on train only.
    encoder, scaler, target_scaler = fit_preprocessors(train_df, config)

    # Persist everything for later training/testing/prediction.
    save_splits_and_artifacts(
        train_df,
        valid_df,
        test_df,
        encoder,
        scaler,
        target_scaler,
        config,
    )


if __name__ == "__main__":
    main()

# Split logic for data separation
# This module handles the splitting of data into training(70%), validation(15%), and test sets(15%)
